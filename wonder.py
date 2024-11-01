import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Sample data loading
data = [
    # Add your scraped data here, e.g.:
    ['02:30', '29-09-2024', 'English Premier League', 'Wolves', 'Liverpool', 1, 2, 3.2, 3.4, 1.4, 2.8, -0.215384615, 0.965517241, 0.523076923, 2.206896552, 0.048076923, 0.202839757, 0.077729387, 0.062088254, 0.153846154, 0.689655172, 0.248451997, 0.211100063, '6.5/4.5/1.45', 4.15, '8.0/5.75/1.33', '45+1', '56', '61', 'NA', 'NA', 'NA', 'NA', 'NA', 'NA', 'NA', 1.60, 9.090277778, '', '', ''],
    ['02:30', '29-09-2024', 'Spanish La Liga', 'Real Sociedad', 'Valencia', 3, 0, 1, 1.6, 0.2, 0.6, -0.231213873, 0.076190476, 0.924855491, 0.19047619, 0.578034682, 0.119047619, 0.6, 0.251775288, 0.578034682, 0.19047619, 0.6, 0.402840461, '1.73/3.4/5.25', 3.46, '1.8/3.25/5.0', '8', '80', '90+1', 'NA', 'NA', 'NA', 'NA', 'NA', 'NA', 'NA', 2.35, 9.202426471, '', '', ''],
    ['02:30', '29-09-2024', 'German Bundesliga', 'Bayern Munchen', 'Bayer Leverkusen', 1, 1, 6, 4.4, 5, 3, 1, -0.6060606, 2.2, 1.818181818, 0.083333333, 0.068870523, 0.034644822, 0.041586341, 0.5, 0.303030303, 0.207868933, 0.182979898, '2.0/3.8/3.3', 3.03333, '1.65/4.5/4.5', '31', '39', 'NA', 'NA', 'NA', 'NA', 'NA', 'NA', 'NA', 'NA', 2.25, 7.288590658, '', '', ''],
    ['02:30', '29-09-2024', 'Italian Serie A', 'Genoa', 'Juventus', 0, 3, 1.8, 1.4, 0.6, 1.2, -0.1333333, 0.324324324, 0.31111111, 0.972972973, 0.123456789, 0.386100386, 0.138028887, 0.535081074, 0.2222222, 0.540540541, 0.248451997, 0.749113504, '4.5/3.3/1.85', 3.21667, '5.25/3.3/1.8', '48', '55', '89', 'NA', 'NA', 'NA', 'NA', 'NA', 'NA', 'NA', 2.25, 7.803414924, '', '', ''],
    # Add more rows as needed...
]

# Define header
header = [
    'Time', 'Date', 'League', 'Home', 'Away', 'GH', 'GA', 'AVGH', 'AVGA', 
    'AGH', 'AGA', 'HGD', 'AGD', 'Goal Value A', 'Goal Value H', 
    'Goal Cost Home', 'Goal Cost Away', 'CV Home', 'CV Away', 
    'Prob.home', 'Prob.Away', 'SD Home', 'SD Away', 'Early', 
    'SUM', 'Live', 'goal1', 'goal2', 'goal3', 'goal4', 
    'goal5', 'goal6', 'goal7', 'goal8', 'goal9', 'goal10', 
    'Over', 'STAKE POOL', 'Team', 'Score', 'W/L'
]

# Create DataFrame
df = pd.DataFrame(data)

# Check if the number of columns matches the header length
if df.shape[1] != len(header):
    raise ValueError(f"Number of columns in data ({df.shape[1]}) does not match header length ({len(header)}).")

# Add header as the first row
df.columns = header

# Helper function to compare values to two decimal places
def is_equal_with_tolerance(value1, value2, tolerance=0.0001):
    return abs(value1 - value2) < tolerance

# Check consecutive identical values function
def check_consecutive_identical_values(column_idx, length=7):
	identical_counts = 0
	prev_value = None
	for row in range(2, len(df) + 2):
		cell_value = worksheet.cell(row=row, column=column_idx).value
		if cell_value == prev_value:
			identical_counts += 1
		else:
			identical_counts = 1  # Reset count if value changes
		if identical_counts >= length:
			return True  # Found at least 7 consecutive identical values
		prev_value = cell_value
	return False

# Save to Excel with formatting
output_file = 'output.xlsx'
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name='Data')

    # Get the workbook and sheet
    workbook = writer.book
    worksheet = writer.sheets['Data']

    # Define fills for background colors	step 1
    fills1 = {
		"dark_purple" : PatternFill(start_color='403151', end_color='403151', fill_type='solid'),  # Dark Purple
		"tan" : PatternFill(start_color='E5B577', end_color='E5B577', fill_type='solid'),          # Tan
		"light_gray" : PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),    # Light Gray
		"early_sky_blue" : PatternFill(start_color='92CDDC', end_color='92CDDC', fill_type='solid'), # Sky Blue
		"early_light_turquoise" : PatternFill(start_color='85DFFF', end_color='85DFFF', fill_type='solid'), # Light Turquoise
		"early_light_yellow" : PatternFill(start_color='FFFF66', end_color='FFFF66', fill_type='solid'), # Light Yellow
		"early_lime" : PatternFill(start_color='CCFF66', end_color='CCFF66', fill_type='solid'),     # Lime
		"early_black" : PatternFill(start_color='000000', end_color='000000', fill_type='solid'),    # Black
		"early_red" : PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'),      # Red
		"turquoise" : PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'),      # Turquoise
		"rose": PatternFill(start_color='FF7D7D', end_color='FF7D7D', fill_type='solid')
	}

	# Define fills for background colors	step 2
    fills2 = {
		"lavender" : PatternFill(start_color='AC75D5', end_color='AC75D5', fill_type='solid'),
		"gold" : PatternFill(start_color='FFD347', end_color='FFD347', fill_type='solid'),
		"light_yellow" : PatternFill(start_color='FFFF66', end_color='FFFF66', fill_type='solid'),
		"red": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
		"black_border" : Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'), top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000')),
		"pink_border" : Border(left=Side(style='thin', color='FF0066'), right=Side(style='thin', color='FF0066'), top=Side(style='thin', color='FF0066'), bottom=Side(style='thin', color='FF0066'))
	}

	# Define fills for background colors	step 3
    fills3 = {
		"purple": PatternFill(start_color="FF00FF", end_color="FF00FF", fill_type="solid"),
		"lavender": PatternFill(start_color="FFAFFF", end_color="FFAFFF", fill_type="solid"),
		"dark_gray": PatternFill(start_color="404040", end_color="404040", fill_type="solid"),
		"light_yellow": PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid"),
		"pink": PatternFill(start_color="FF0066", end_color="FF0066", fill_type="solid"),
		"dark_purple": PatternFill(start_color="5B114D", end_color="5B114D", fill_type="solid"),
		"light_green": PatternFill(start_color="7DFF7D", end_color="7DFF7D", fill_type="solid"),
		"aqua": PatternFill(start_color="00FFFF", end_color="00FFFF", fill_type="solid"),
		"gray": PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid"),
		"light_orange": PatternFill(start_color="FABF8F", end_color="FABF8F", fill_type="solid"),
		"blue": PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid"),
		"sky_blue": PatternFill(start_color="92CDDC", end_color="92CDDC", fill_type="solid"),
		"red": PatternFill(start_color="FF3333", end_color="FF3333", fill_type="solid"),
		"dark_red": PatternFill(start_color="A50021", end_color="A50021", fill_type="solid"),
		"gold": PatternFill(start_color="C8A200", end_color="C8A200", fill_type="solid")
	}

	# Define fills for background colors	step 4
    fills4 = {
        "teal": PatternFill(start_color='33CCCC', end_color='33CCCC', fill_type='solid'),
        "red": PatternFill(start_color='FF3333', end_color='FF3333', fill_type='solid'),
        "gold": PatternFill(start_color='FACA00', end_color='FACA00', fill_type='solid'),
        "light_yellow": PatternFill(start_color='FFFF66', end_color='FFFF66', fill_type='solid'),
        "sky_blue": PatternFill(start_color='92CDDC', end_color='92CDDC', fill_type='solid'),
        "turquoise": PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid'),
        "pink": PatternFill(start_color='FF0066', end_color='FF0066', fill_type='solid'),
        "blue": PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid'),
        "black": PatternFill(start_color='000000', end_color='000000', fill_type='solid'),
        "aqua": PatternFill(start_color='00FFFF', end_color='00FFFF', fill_type='solid'),
        "white": PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid'),
        "gray": PatternFill(start_color='A6A6A6', end_color='A6A6A6', fill_type='solid'),
        "light_green": PatternFill(start_color='ABFFDD', end_color='ABFFDD', fill_type='solid'),
        "lavender": PatternFill(start_color='AC75D5', end_color='AC75D5', fill_type='solid'),
        "rose": PatternFill(start_color='DA9694', end_color='DA9694', fill_type='solid'),
        "light_turquoise": PatternFill(start_color='85DFFF', end_color='85DFFF', fill_type='solid'),
        "dark_purple": PatternFill(start_color='403151', end_color='403151', fill_type='solid'),
        "dark_red": PatternFill(start_color='A50021', end_color='A50021', fill_type='solid'),
        "green": PatternFill(start_color='33CC33', end_color='33CC33', fill_type='solid')
    }

    # Define border styles
    borders4 = {
        "black": Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'), top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000')),
        "blue": Border(left=Side(style='thin', color='0000FF'), right=Side(style='thin', color='0000FF'), top=Side(style='thin', color='0000FF'), bottom=Side(style='thin', color='0000FF')),
        "pink": Border(left=Side(style='thin', color='FF0066'), right=Side(style='thin', color='FF0066'), top=Side(style='thin', color='FF0066'), bottom=Side(style='thin', color='FF0066'))
    }

    # Font styles
    font_red_bold_italic = Font(color="FF0000", bold=True, italic=True)
    font_purple_bold_italic = Font(color="FF00FF", bold=True, italic=True)
    font_bold_italic_arial_narrow = Font(name='Arial Narrow', size=11, bold=True, italic=True)
    font_gold_bold_italic = Font(bold=True, italic=True, color='FFD700')
    font_blue_bold = Font(bold=True, color='0000FF')
    font_aqua_bold_italic	= Font(bold=True, italic=True, color='00FFFF')
    font_blue_bold_italic = Font(name='Arial', size=11, color='0000FF', bold=True, italic=True)

	# Font styles
    font_bold_italic = Font(bold=True, italic=True)
    font_blue = Font(color='0000FF')
    font_red = Font(color='FF0000')
    font_white = Font(color='FFFFFF', bold=True, italic=True)
    font_black = Font(color='000000')
    font_aqua = Font(color='00FFFF', bold=True, italic=True)

    font_gold = Font(color='FACA00', bold=True, italic=True)
    font_yellow = Font(color='FFFF00', bold=True, italic=True)
    font_gray = Font(color='A6A6A6')

    # Alignment
    alignment_center = Alignment(horizontal='center', vertical='center')

    # Apply alignment, column width adjustments, and highlight formatting
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter  # Get the column letter
        for cell in column:
            cell.alignment = alignment_center
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        worksheet.column_dimensions[column_letter].width = adjusted_width

    # Apply Arial Narrow formatting to Column E (Away)
    for cell in worksheet['E']:
        cell.font = font_bold_italic_arial_narrow  # Set font style for the entire column

    # Iterate through rows and apply conditional formatting
    for row in range(2, len(df) + 2):  # Start from 2 to skip header
        prob_home = worksheet.cell(row=row, column=20).value
        prob_away = worksheet.cell(row=row, column=21).value
        hgd = worksheet.cell(row=row, column=12)
        agd = worksheet.cell(row=row, column=13)
        avg = (worksheet.cell(row=row, column=8).value + worksheet.cell(row=row, column=9).value) / 2
        ags = (worksheet.cell(row=row, column=10).value + worksheet.cell(row=row, column=11).value) / 2
        sd_home = worksheet.cell(row=row, column=22).value
        sd_away = worksheet.cell(row=row, column=23).value
        sd = sd_home + sd_away
        cell_early = worksheet.cell(row=row, column=24)
        early_value = cell_early.value
        draw_odd = float(early_value.split('/')[1].strip())  # Assuming 'Early' is in the 24th column

        goal_value_away = worksheet.cell(row=row, column=14)
        goal_value_home = worksheet.cell(row=row, column=15)

        goal_cost_home = worksheet.cell(row=row, column=16).value
        goal_cost_away = worksheet.cell(row=row, column=17).value
        cv_home = worksheet.cell(row=row, column=18).value
        cv_away = worksheet.cell(row=row, column=19).value
        sum_sd = (sd_home or 0) + (sd_away or 0)
        sum_cv = (cv_home or 0) + (cv_away or 0)
        live_cell = worksheet.cell(row=row, column=26)
        sum_cell = worksheet.cell(row=row, column=25)
        
		# Check if SCR or ECC is in any of the goal columns
        is_cancelled = any(
            worksheet.cell(row=row, column=col).value in ['SCR', 'ECC'] for col in range(27, 31)
        )

        # DARK PURPLE highlight in Early column
        if (1 / 1.57 <= prob_home <= 1 / 1.01 or 1 / 1.57 <= prob_away <= 1 / 1.01) and (avg > ags) and (ags > sd):
            worksheet.cell(row=row, column=24).fill = fills1["rose"]  # Highlight EARLY column in ROSE
        
        if (1 / 2.63 <= prob_home <= 1 / 1.01 or 1 / 2.63 <= prob_away <= 1 / 1.01) and (avg > ags) and (ags > sd):
            worksheet.cell(row=row, column=24).fill = fills1["early_sky_blue"]  # Highlight EARLY column in SKY BLUE

        if (1 / 2.63 <= prob_home <= 1 / 1.01 or 1 / 2.63 <= prob_away <= 1 / 1.01) and (avg < ags) and (ags > sd):
            worksheet.cell(row=row, column=24).fill = fills1["early_light_turquoise"]  # Highlight EARLY column in LIGHT TURQUOISE

        if not is_cancelled:
            # Zoya - Lime Highlight
            if agd == 0 and (prob_away > avg < sd <= 0.81 or prob_home > avg < sd <= 0.81):
                cell_early.fill = fills4["early_lime"]

            elif agd == 0 and (prob_away > avg > sd <= 0.81 or prob_home > avg > sd <= 0.81):
                cell_early.fill = fills1["early_black"]

            # Zoya - Red Highlight
            elif avg < ags <= sd <= 0.71:
                cell_early.fill = fills1["early_red"]

            # Zoya - Turquoise Highlight
            elif avg >= ags > sd <= 0.71:
                cell_early.fill = fills1["turquoise"]
		
        if (1 / 1.6 >= prob_home >= 1 / 2.63 or 1 / 1.6 >= prob_away >= 1 / 2.63) and avg > ags < sd and sd <= 0.71 + draw_odd >= 3.40:
            worksheet.cell(row=row, column=24).fill = fills1["dark_purple"]

        # TAN highlight in Early column
        elif (1 / 1.01 >= prob_home >= 1 / 2.63 or 1 / 1.01 >= prob_away >= 1 / 2.63) and avg < ags < sd and sd <= 0.57 + draw_odd >= 3.40:
            worksheet.cell(row=row, column=24).fill = fills1["tan"]

        # LIGHT GRAY highlight in Early column
        elif (1 / 1.60 >= prob_home >= 1 / 2.63 or 1 / 1.60 >= prob_away >= 1 / 2.63) and avg > ags > sd and sd <= 0.57 + draw_odd >= 3.40:
            worksheet.cell(row=row, column=24).fill = fills1["light_gray"]


        # Set red font style for 'Early' if specific conditions are met
        if 'SCR' in df.columns.values or 'ECC' in df.columns.values:
            if any(cell.value in ('SCR', 'ECC') for cell in worksheet[row]):
                worksheet.cell(row=row, column=24).font = font_red_bold_italic

        # Set purple font style for 'SUM' if specific conditions are met
        if 'ECC' in df.columns.values:
            worksheet.cell(row=row, column=25).font = font_purple_bold_italic

		# LIVE Column conditions for SUM of SD
        if 0 <= sum_sd <= 0.27:
            live_cell.fill = fills2["lavender"]
        elif 0.28 <= sum_sd <= 0.44:
            live_cell.fill = fills4["red"]
        elif 0.45 <= sum_sd <= 0.60:
            live_cell.fill = fills2["gold"]
        elif 0.61 <= sum_sd <= 0.81:
            live_cell.fill = fills2["light_yellow"]

        highest_prob = max(prob_home, prob_away)

        # SUM Column conditions for PROB + Goal Cost + CV sum
        print(highest_prob + goal_cost_home + sum_cv)
        if 0 <= highest_prob + goal_cost_home + sum_cv <= 0.17:
            sum_cell.fill = fills2["lavender"]
        elif 0.18 <= highest_prob + goal_cost_home + sum_cv <= 0.26:
            sum_cell.fill = fills2["red"]
        elif 0.27 <= highest_prob + goal_cost_home + sum_cv <= 0.40:
            sum_cell.fill = fills2["gold"]
        elif 0.41 <= highest_prob + goal_cost_home + sum_cv <= 0.91:
            sum_cell.fill = fills2["light_yellow"]

        # F-09: Border conditions based on Zoya conditions (adjusted logic with constraints)
        zoya_values = ['Rose', 'Light Yellow', 'Red', 'Light Gray', 'Dark Purple', 'Sky Blue', 'Light Turquoise', 'Turquoise', 'Tan']
        if worksheet.cell(row=row, column=25).value in zoya_values and 0.11 <= min(sd_home, sd_away) <= 0.31:
            if min(sd_home, sd_away) * 1.25 <= max(sd_home, sd_away) * 1.17:
                if worksheet.cell(row=row, column=25).value in ['Rose', 'Light Yellow', 'Red', 'Light Gray', 'Dark Purple']:
                    live_cell.border = fill_black_border2
                elif worksheet.cell(row=row, column=25).value in ['Sky Blue', 'Light Turquoise', 'Turquoise', 'Tan']:
                    live_cell.border = fill_pink_border2
		
		#========================== Step 3 Conditions ============================
		# PAV
        if is_equal_with_tolerance(goal_cost_home * 1.04, hgd.value):
            hgd.fill = fills3["purple"]
        if is_equal_with_tolerance(goal_cost_away * 1.04, agd.value):
            agd.fill = fills3["purple"]

		# HQO
        if is_equal_with_tolerance(goal_cost_home * 0.96, hgd.value):
            hgd.fill = fills3["lavender"]
        if is_equal_with_tolerance(goal_cost_away * 0.96, agd.value):
            agd.fill = fills3["lavender"]

		# CIS (CEO)
        if 1.4 <= goal_cost_home <= 1.51 and is_equal_with_tolerance(goal_cost_home, hgd.value):
            hgd.fill = fills3["dark_gray"]
        if 1.4 <= goal_cost_away <= 1.51 and is_equal_with_tolerance(goal_cost_away, agd.value):
            agd.fill = fills3["dark_gray"]

		# ASY (PAV)
        if 1.4 <= hgd.value <= 1.51 and is_equal_with_tolerance(goal_cost_home, hgd.value):
            hgd.fill = fills3["light_yellow"]
        if 1.4 <= agd.value <= 1.51 and is_equal_with_tolerance(goal_cost_away, agd.value):
            agd.fill = fills3["light_yellow"]

		# CEO (PAV)
        if is_equal_with_tolerance(goal_cost_home * 2.8, hgd.value):
            hgd.fill = fills3["pink"]
        if is_equal_with_tolerance(goal_cost_away * 2.8, agd.value):
            agd.fill = fills3["pink"]

		# AAO (CEO)
        if is_equal_with_tolerance(hgd.value * 2.8, goal_cost_home):
            hgd.fill = fills3["dark_purple"]
        if is_equal_with_tolerance(agd.value * 2.8, goal_cost_away):
            agd.fill = fills3["dark_purple"]

		# COO (PAV)
        if is_equal_with_tolerance(goal_cost_home * 2.88, hgd.value):
            hgd.fill = fills3["light_green"]
        if is_equal_with_tolerance(goal_cost_away * 2.88, agd.value):
            agd.fill = fills3["light_green"]

		# THU (PAV)
        if is_equal_with_tolerance(goal_cost_home * 3.0, hgd.value):
            hgd.fill = fills3["aqua"]
        if is_equal_with_tolerance(goal_cost_away * 3.0, agd.value):
            agd.fill = fills3["aqua"]

		# CIN (PAV)
        if is_equal_with_tolerance(goal_cost_home * 0.96, goal_value_away.value):
            goal_value_away.fill = fills3["pink"]
        if is_equal_with_tolerance(goal_cost_away * 0.96, goal_value_home.value):
            goal_value_home.fill = fills3["pink"]

		# HSS (PAV)
        if is_equal_with_tolerance(goal_cost_home * 1.04, goal_value_away.value):
            goal_value_away.fill = fills3["lavender"]
        if is_equal_with_tolerance(goal_cost_away * 1.04, goal_value_home.value):
            goal_value_home.fill = fills3["lavender"]

		# HOY (VLN)
        if is_equal_with_tolerance(goal_cost_home * 1.12, goal_value_away.value):
            goal_value_away.fill = fills3["gray"]
        if is_equal_with_tolerance(goal_cost_away * 1.12, goal_value_home.value):
            goal_value_home.fill = fills3["gray"]

		# WHP (VLN)
        if is_equal_with_tolerance(goal_cost_home * 0.88, goal_value_away.value):
            goal_value_away.fill = fills3["light_orange"]
        if is_equal_with_tolerance(goal_cost_away * 0.88, goal_value_home.value):
            goal_value_home.fill = fills3["light_orange"]

		# VLN (PAV)
        if is_equal_with_tolerance(goal_cost_home * 1.12, hgd.value):
            hgd.fill = fills3["blue"]
        if is_equal_with_tolerance(goal_cost_away * 1.12, agd.value):
            agd.fill = fills3["blue"]

		# VLN (HSY)
        if is_equal_with_tolerance(goal_cost_home * 0.88, hgd.value):
            hgd.fill = fills3["sky_blue"]
        if is_equal_with_tolerance(goal_cost_away * 0.88, agd.value):
            agd.fill = fills3["sky_blue"]

		# HOC
        if is_equal_with_tolerance(goal_cost_home * 2 * 1.04, hgd.value):
            hgd.fill = fills3["red"]
        if is_equal_with_tolerance(goal_cost_away * 2 * 1.04, agd.value):
            agd.fill = fills3["red"]

		# HSY
        if is_equal_with_tolerance(goal_cost_home * 2 * 0.96, hgd.value):
            hgd.fill = fills3["blue"]
        if is_equal_with_tolerance(goal_cost_away * 2 * 0.96, agd.value):
            agd.fill = fills3["blue"]

		# OFF (HOC)
        if is_equal_with_tolerance(goal_cost_home * 2 * 1.12, hgd.value):
            hgd.fill = fills3["dark_red"]
        if is_equal_with_tolerance(goal_cost_away * 2 * 1.12, agd.value):
            agd.fill = fills3["dark_red"]
        if is_equal_with_tolerance(goal_cost_home * 2 * 0.88, hgd.value):
            hgd.fill = fills3["gold"]
        if is_equal_with_tolerance(goal_cost_away * 2 * 0.88, agd.value):
            agd.fill = fills3["gold"]

		#========================== Step 4=======================
		 # LIVE Column conditions for SUM of SD
        if 0 <= sum_sd <= 0.27:
            live_cell.fill = fills4["lavender"]
        elif 0.28 <= sum_sd <= 0.44:
            live_cell.fill = fills4["red"]
        elif 0.45 <= sum_sd <= 0.60:
            live_cell.fill = fills4["gold"]
        elif 0.61 <= sum_sd <= 0.81:
            live_cell.fill = fills4["light_yellow"]

        # 3DD conditions for SUM Column
        if 0 <= prob_home + goal_cost_home + sum_cv <= 0.17:
            sum_cell.fill = fills4["lavender"]
        elif 0.18 <= prob_home + goal_cost_home + sum_cv <= 0.26:
            sum_cell.fill = fills4["red"]
        elif 0.27 <= prob_home + goal_cost_home + sum_cv <= 0.40:
            sum_cell.fill = fills4["gold"]
        elif 0.41 <= prob_home + goal_cost_home + sum_cv <= 0.91:
            sum_cell.fill = fills4["light_yellow"]

        # F-09: Border conditions based on Zoya conditions
        zoya_values = ['Rose', 'Light Yellow', 'Red', 'Light Gray', 'Dark Purple', 'Sky Blue', 'Light Turquoise', 'Turquoise', 'Tan']
        lowest_sd = min(sd_home, sd_away)
        if worksheet.cell(row=row, column=25).value in zoya_values and 0.11 <= min(sd_home, sd_away) <= 0.31:
            if min(sd_home, sd_away) * 1.25 <= max(sd_home, sd_away) * 1.17:
                if worksheet.cell(row=row, column=25).value in ['Rose', 'Light Yellow', 'Red', 'Light Gray', 'Dark Purple']:
                    live_cell.border = Border(left=Side(style='thick', color='000000'))  # Black border
                elif worksheet.cell(row=row, column=25).value in ['Sky Blue', 'Light Turquoise', 'Turquoise', 'Tan']:
                    live_cell.border = Border(left=Side(style='thick', color='FF0066'))  # Pink border

        # RV conditions for Goal Value Home and Away
        if row >= 8:  # Ensure at least 7 previous rows to check
            hgd_values = [worksheet.cell(r, 11).value for r in range(row - 7, row)]
            agd_values = [worksheet.cell(r, 12).value for r in range(row - 7, row)]
            if len(set(hgd_values)) == 1 and hgd_values[0] == worksheet.cell(row, 15).value:  # Check for HGD
                worksheet.cell(row, 15).fill = fills4["teal"]  # Goal Value Home
            if len(set(agd_values)) == 1 and agd_values[0] == worksheet.cell(row, 14).value:  # Check for AGD
                worksheet.cell(row, 14).fill = fills4["teal"]  # Goal Value Away

        # D3rd conditions for GH column
        if row >= 4:  # Ensure at least 3 previous matches to check
            previous_matches = [(worksheet.cell(row - i, 6).value, worksheet.cell(row - i, 7).value) for i in range(1, 4)]
            draw_match_found = any(match == (2, 2) or match == (3, 3) for match in previous_matches)
            if draw_match_found:
                worksheet.cell(row, 6).font = font_blue_bold

        # D5th conditions for Time column
        if row >= 6:  # Ensure at least 5 previous matches to check
            previous_matches = [(worksheet.cell(row - i, 6).value, worksheet.cell(row - i, 7).value) for i in range(1, 6)]
            draw_match_found = any(match == (2, 2) or match == (3, 3) for match in previous_matches)
            if draw_match_found:
                worksheet.cell(row, 1).fill = fills4["sky_blue"]

        # HDN conditions for AVGA column
        if row >= 6:  # Ensure at least 5 previous matches to check
            previous_matches = [(worksheet.cell(row - i, 6).value, worksheet.cell(row - i, 7).value) for i in range(1, 6)]
            draw_match_found = any(match == (0, 0) or match == (1, 1) for match in previous_matches)
            if draw_match_found:
                worksheet.cell(row, 22).font = font_gold_bold_italic

        # D1R conditions for AVGH column
        #if row >= 2:  # Check only the last match
          #  last_match = (worksheet.cell(row - 1, 6).value, worksheet.cell(row - 1, 7).value)
          #  draw_match_found = last_match in [(2, 2), (3, 3)] or (last_match[0] >= 4 or last_match[1] >= 4)
          #  if draw_match_found:
          #      worksheet.cell(row, 21).font = font_styles["yellow_bold_italic"]

        # D77 conditions for League column
        if row >= 6:  # Ensure at least 5 previous matches to check
            previous_matches = [(worksheet.cell(row - i, 6).value, worksheet.cell(row - i, 7).value) for i in range(1, 6)]
            goals_found = any(sum(match) >= 5 for match in previous_matches)
            if goals_found:
                worksheet.cell(row, 3).border = Border(left=Side(style='thick', color='0000FF'))  # Blue border

        # NK conditions for Away column
        # Assuming values from previous conditions will affect the NK condition
        nk_condition = (
            (worksheet.cell(row, 26).fill == fills4["pink"] or  # If LIVE has pink border
            worksheet.cell(row, 25).fill == fills4["red"] or  # If SUM has red fill
            worksheet.cell(row, 21).font.color == 'FFFF00' or  # If AVGH has yellow font
            worksheet.cell(row, 6).font.color == 'FF0000' or  # If GH has red font
            worksheet.cell(row, 1).fill == fills4["sky_blue"])  # If Time has sky blue fill
        )
        if nk_condition:
            worksheet.cell(row, 5).border = Border(left=Side(style='thick', color='FF0000'))  # Red border

        # IMC conditions for Goal Value Home column
        imc_condition = (
            (worksheet.cell(row, 15).fill == fills4["pink"] or  # If Goal Value Home has pink fill
            worksheet.cell(row, 14).font.color == 'FF0000' or  # If Goal Value Away has red font
            worksheet.cell(row, 21).font.color == 'FFFF00')  # If AVGH has yellow font
        )
        if imc_condition:
            worksheet.cell(row, 15).border = Border(left=Side(style='thick', color='FF0066'))  # Pink border

        # LIT conditions for AGD column
        if worksheet.cell(row, 12).value == 0 and (sd_home + sd_away) >= 1.79:
            worksheet.cell(row, 12).fill = fills4["turquoise"]  # Highlight AGD

        # TX conditions for Live column
        tx_condition = (
            (worksheet.cell(row, 15).value + worksheet.cell(row, 16).value >= 2.2)  # Check TCC condition
        )
        if tx_condition:
            worksheet.cell(row, 26).font = font_aqua_bold_italic  # Color Live font

        if worksheet.cell(row=row, column=3).value == 'TCC' or worksheet.cell(row=row, column=3).value == 'ECC':
            sum_goal_value = worksheet.cell(row=row, column=14).value + worksheet.cell(row=row, column=13).value
            if sum_goal_value >= 1.81 and len(set(zoya_values) & set(worksheet.cell(row=row, column=25).value.split(','))) >= 2:
                live_cell.border = fills4["black_border"]

        # NN condition
        if worksheet.cell(row=row, column=3).value == 'TCC' and worksheet.cell(row=row, column=12).value == 'LN-7':
            worksheet.cell(row=row, column=12).font = font_bold_italic

        # NED condition
        if worksheet.cell(row=row, column=12).value == 'NN' and worksheet.cell(row=row, column=3).value == 'TIG':
            worksheet.cell(row=row, column=3).font = font_yellow_bold

        # NFD condition
        if worksheet.cell(row=row, column=3).value == 'TCC' and (
                worksheet.cell(row=row, column=26).value in ['F-09', 'Nil to Nil']):
            worksheet.cell(row=row, column=21).border = fills4["red"]

        # FTF condition
        if (worksheet.cell(row=row, column=26).value in ['F-09', 'DBT', 'Nil to Nil'] and
                worksheet.cell(row=row, column=20).value == 'AH' and
                worksheet.cell(row=row, column=12).value in ['LN-7', 'ATOL']):
            worksheet.cell(row=row, column=22).border = fills4["pink_border"]

        # EVN condition
        if (worksheet.cell(row=row, column=24).value == 'HDN' and
                worksheet.cell(row=row, column=20).value in ['AH', 'SK', 'BK', 'NX', 'F-09', 'Nil to Nil'] and
                (worksheet.cell(row=row, column=23).value in ['3NR', 'LN-7', 'Double CW'])):
            worksheet.cell(row=row, column=23).border = fills4["red"]

        # TIG condition
        if (worksheet.cell(row=row, column=3).value in ['TCC', 'ECC'] and
			(worksheet.cell(row=row, column=20).value in ['AH', 'Double CW']) and
			(worksheet.cell(row=row, column=21).value in ['3NR', 'LN-7', 'TX']) and
			(sum_cv <= 0.15)):
			# Your code logic here

            worksheet.cell(row=row, column=24).border = fills4["black_border"]

        # EAG condition
        if (worksheet.cell(row=row, column=26).value in ['3DD', 'DP'] and
                (worksheet.cell(row=row, column=3).value in ['IMC', 'AH', 'SK']) and
                (worksheet.cell(row=row, column=3).value in ['TIG', 'DP', 'ZION'])):
            worksheet.cell(row=row, column=4).border = fills4["blue"]

        # CCH condition
        if (worksheet.cell(row=row, column=3).value not in ['TCC'] and
                worksheet.cell(row=row, column=3).value == 'ECC' and
                worksheet.cell(row=row, column=26).value in ['F-09', 'Nil to Nil']):
            worksheet.cell(row=row, column=2).font = font_bold_italic

        # CYC condition
        if (worksheet.cell(row=row, column=3).value in ['IMC', 'TIG', 'EAG']):
            worksheet.cell(row=row, column=5).font = font_sky_blue

        # OWO condition
        if (worksheet.cell(row=row, column=3).value in ['TCC', 'ECC', 'NK'] and
                worksheet.cell(row=row, column=26).value in ['F-09', 'FTF'] and
                worksheet.cell(row=row, column=21).value in ['3NR', 'LN-7']):
            worksheet.cell(row=row, column=6).border = fills4["red"]

        # BFM condition
        if (worksheet.cell(row=row, column=3).value in ['TCC', 'ECC', 'AH', 'BK'] and
                worksheet.cell(row=row, column=10).value in ['D1R', 'HDN', 'GAB']):
            worksheet.cell(row=row, column=7).border = fills4["black"]

        # OHO condition
        if (worksheet.cell(row=row, column=3).value not in ['ECC', 'NX'] and
                worksheet.cell(row=row, column=3).value in ['TCC', 'AH'] and
                (worksheet.cell(row=row, column=20).value in ['SK', 'TIG']) and
                sum_goal_value >= 1.89):
            worksheet.cell(row=row, column=8).font = Font(color='A6A6A6')

        # OAO condition
        if (worksheet.cell(row=row, column=3).value in ['3DD', 'DP', 'AH', 'BAG'] and
                sum_goal_value >= 2.1):
            worksheet.cell(row=row, column=22).border = fills4["gold"]

        # VNM condition
        if (worksheet.cell(row=row, column=3).value in ['3DD', 'DP', 'NN', 'NED'] and
                sum_goal_value >= 1.83):
            worksheet.cell(row=row, column=22).fill = fills4["turquoise"]

        # HHY condition
        if (worksheet.cell(row=row, column=3).value == 'TCC' and
                (worksheet.cell(row=row, column=10).value in ['HDN', 'D1R', 'SK', 'AH']) and
                sum_goal_value >= 1.95):
            worksheet.cell(row=row, column=24).border = fills4["blue"]

        # SYM condition
        if (worksheet.cell(row=row, column=3).value == 'TIG' and
                sum_goal_value >= 1.74):
            worksheet.cell(row=row, column=24).fill = fills4["light_green"]

        # ICE condition
        three_sequence_numbers = [0.314678903, 0.321265432, 0.338775623]
        if len(set(three_sequence_numbers)) >= 3:
            smallest_number = min(three_sequence_numbers)
            # Apply formatting to smallest number if found
            # Note: Find the column index for the smallest number and apply AQUA fill
            for col_idx in range(1, len(header) + 1):
                if worksheet.cell(row=row, column=col_idx).value == smallest_number:
                    worksheet.cell(row=row, column=col_idx).fill = fills4["aqua"]
                    break

        # GD's-Z condition
        if worksheet.cell(row=row, column=14).value == 0 and sum_goal_value >= 2.6:
            worksheet.cell(row=row, column=4).border = fills4["purple"]

	#================== Step 5 =====================
	# TWI and PTN conditions
        if hgd and goal_value_away and str(hgd.value).split(".")[1][:2] == str(goal_value_away.value).split(".")[1][:2]:
            worksheet.cell(row=row, column=15).fill = fills4["rose"]
        if agd and goal_value_home and str(agd).split(".")[1][:2] == str(goal_value_home.value).split(".")[1][:2]:
            worksheet.cell(row=row, column=15).fill = fills4["rose"]

        # PTN: Border of Goal Value Away in Blue
        if hgd and agd and str(hgd.value).split(".")[1][:2] == str(agd).split(".")[1][:2]:
            worksheet.cell(row=row, column=14).border = borders4["blue"]

        # SK conditions
        if prob_home and goal_value_away and str(prob_home).split(".")[1][:8] == str(goal_value_away.value).split(".")[1][:8] and prob_home <= 0.2:
            worksheet.cell(row=row, column=21).fill = fills4["light_turquoise"]

        # NX conditions (similar to SK)
        if prob_home and hgd and str(prob_home).split(".")[1][:7] == str(hgd.value).split(".")[1][:7]:
            worksheet.cell(row=row, column=23).fill = fills4["dark_purple"]

        # AH conditions (same formula as RV)
        if hgd and prob_home and str(hgd.value).split(".")[1][:7] == str(prob_home).split(".")[1][:7]:
            worksheet.cell(row=row, column=22).border = borders4["blue"]

        # BNG and HCO conditions
        if goal_cost_home and goal_value_away and abs(goal_cost_home * 2 * 1.04 - goal_value_away.value) < 0.01:
            worksheet.cell(row=row, column=12).fill = fills4["light_green"]
        if goal_cost_away and goal_value_home and abs(goal_cost_away * 2 * 1.04 - goal_value_home.value) < 0.01:
            worksheet.cell(row=row, column=13).fill = fills4["light_green"]

        # BK09 conditions
        if hgd and goal_cost_home and abs(hgd.value * 2 * 1.12 - goal_cost_home) < 0.01:
            worksheet.cell(row=row, column=15).fill = fills4["dark_red"]

        # QM conditions
        if hgd and goal_cost_home and abs(hgd.value * 2 * 0.88 - goal_cost_home) < 0.01:
            worksheet.cell(row=row, column=15).fill = fills4["green"]

        # SCA conditions: Highlight font in BLUE for "Live" when Zoya conditions are met
        zoya_terms = ['PAV', 'HQO', 'VLN', 'HOC', 'HCO', 'HSY', 'GDZ', 'PTN', 'TWIs', 'BNG', 'AH', 'NX', 'UD', 'SK']
        if any(term in str(worksheet.cell(row=row, column=25).value) for term in zoya_terms):
            worksheet.cell(row=row, column=25).font = font_blue_bold_italic

# Done
print(f"Data with headers and formatting saved to {output_file}.")
